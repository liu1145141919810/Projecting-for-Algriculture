import pandas as pd
import openpyxl
import warnings
from openpyxl.utils.cell import column_index_from_string
df1=pd.read_excel('file1.xlsx')
df22=pd.read_excel('file2.xlsx',sheet_name=None)#读取文件
boxbox1=[]
boxbox2=[]
i=0
for key in df22.keys():
    i+=1
    if i==1:boxbox1.append(key)
    if i==2:
        boxbox2.append(key)
        break
df2=df22[boxbox1[0]]
df3=df22[boxbox2[0]]
#以上是为了得到目标表格
class Plan:
    #用来快速得到字符串和其在数据集对应位置的方法
    plant_lexico=["粮食","豆类","蔬菜 ","食用菌"]
    blockname_lexico=["平旱地","梯田","山坡地","水浇地","普通大棚 ","智慧大棚"]
    block_reverse={
        "平旱地":0,
        "梯田":1,
        "山坡地":2,
        "水浇地":3,
        "普通大棚 ":4,
        "智慧大棚":5
    }
    #以下四个记录作物属于的大类
    cai_lexico={}
    liang_lexico={}
    jun_lexico={}
    dou_lexico={}
    class block:
        def __init__(self,selfname,product_memo,amount): 
            self.status=[1,0]#表示地块的状态，第一个表示种植的种类，第二个表示种了几个季节 ，初始化为如是
            self.amount=amount#地块的面积
            self.trueamount=amount#计算用数据
            self.nowamount=0#已经种植的面积
            self.selfname=selfname#记录地块种类的名字
            self.memo=product_memo# 记录上一次地块种了什么
            self.newmemo=[[],[]]# 记录这一次地块中了什么
            self.limittimes=1 # 
            self.update()#记录有多久没种豆
            self.result=[[],[]]#打印结果用
            self.recify(self.selfname)#给状态赋值
            ## above is the main part
        ## main use_tool is the next function
        def add_element(self,productname,size):#外部控制保证了输入的size不会超过地块的面积
            #向该地块种植某种作物
            #以下根据不同的状态作运算
            #print(self.status)
            if size<=0.00000001:
                return 0
            size=round(size,4)
            sep_cond1=self.status==[3,1] or self.status==[4,1] or (self.status==[2,1] and productname!="水稻")
            sep_cond2=self.status[0]==1 or (self.status==[2,0] and productname=="水稻")
            sep_cond3=self.status==[4,0] or self.status==[3,0]
            sep_cond4=self.status==[2,0] and productname!="水稻"
            if sep_cond2:
                if self.pretelling(productname):#判断能不能种该种类型
                    self.newmemo[0].append(productname)#今年的记录更新
                    self.result[0].append([productname,size])#结果更新
                    self.nowamount+=size#已种的量更新
                    if 0==self.amount-self.nowamount:#中完了
                        self.status[1]+=1#种的季节加一
                        self.memo=self.newmemo#为下一年记录作准备
                    self.update()#更新豆子的限制
                    return 1    #种植成功
            if sep_cond1:
                if self.pretelling(productname):
                    self.newmemo[1].append(productname)
                    self.result[1].append([productname,size])
                    self.nowamount+=size
                    if 0==self.amount-self.nowamount:
                        self.memo=self.newmemo
                    self.update()
                    return 1    
            if sep_cond3:
                if self.pretelling(productname):
                    self.newmemo[0].append(productname)
                    self.result[0].append([productname,size])
                    self.nowamount+=size
                    if  0==self.amount-self.nowamount:
                        self.update()
                    return 1    
            if  sep_cond4:
                if self.pretelling(productname):
                    self.newmemo[0].append(productname)
                    self.result[0].append([productname,size])
                    self.nowamount+=size
                    if 0==self.amount-self.nowamount*2:
                        self.amount*=2
                        self.status[1]+=1
                        self.update()
                    return 1
            return 0  #种植失败                 
        def recify(self,selfname):#初始状态赋予
            teller0=selfname==Plan.blockname_lexico[0]
            teller0=teller0 or selfname==Plan.blockname_lexico[1]
            teller0=teller0 or selfname==Plan.blockname_lexico[2]
            teller1=selfname==Plan.blockname_lexico[3]
            teller2=selfname==Plan.blockname_lexico[4]
            teller3=selfname==Plan.blockname_lexico[5]
            if teller0:
                self.status=[1,0]
            if teller1:
                self.status=[2,0]
            if teller2:
                self.status=[3,0]
            if teller3:
                self.status=[4,0]
        def update(self):#检测到有豆子，将限制次数清零
            for j in self.memo:
                for i in j:
                    try:
                        value=Plan.dou_lexico[i]
                        if value == 1:
                            self.limittimes=0
                            break
                    except:
                        pass    
        def pretelling(self,productname):#判断能否种某种作物
            if self.status==[1,0]:
                if Plan.liang_lexico[productname]!=1:#不能种粮食以外
                    return False
            if self.status==[1,1]:#种完了
                return False
            if self.status==[2,0]:
                if Plan.cai_lexico[productname]!=1 and productname!="水稻":
                    return False
            if self.status==[2,1]:
                if  self.newmemo[0]==["水稻"]:
                    return False
                else:
                    if Plan.cai_lexico[productname]!=1:
                        return False
            if self.status==[2,2]:
                return False
            if self.status==[3,0]:
                if Plan.jun_lexico[productname]!=1 and Plan.cai_lexico[productname]!=1:
                    return False
            if self.status==[3,1]:
                if Plan.jun_lexico[productname]!=1 and Plan.cai_lexico[productname]!=1:
                    return False
                else:
                    if Plan.cai_lexico[productname]==1:
                        if self.newmemo[0]!=[] and Plan.jun_lexico[self.newmemo[0][0]]!=1:
                            return False
                    if Plan.jun_lexico[productname]==1:
                        if self.newmemo[0]!=[] and Plan.cai_lexico[self.newmemo[0][0]]!=1:
                            return False
            if self.status==[3,2]:
                return False
            if self.status==[4.0] or self.status==[4,1]:
                if Plan.cai_lexico[productname]!=1:
                    return False
            if self.status==[4,2]:
                return False 
            # 以上是在说每种情况不能种的作物，将之排除
            #######
            if self.nowamount>=self.amount:#田种满了
                return False
            if self.limittimes==2 and Plan.dou_lexico[productname]!=1:
                #已经两年了还不中豆
                return False
            for i in self.memo:#和去年有一样的东西
                for j in i:
                    if j==productname:
                        return False
            return True#能中了                          
    class datatype:
        def __init__(self,blockname,low,high,cost,productivity,Teller):
            #计算利润并说明该种田还能不能种
            self.blockname=blockname
            if Teller:            
                benifit=((high+low)/2)*productivity-cost
                self.benifit=benifit
                self.productivity=productivity
                self.allow=True
                self.prepare=benifit
                self.i=1
            else:
                self.i=0
                self.benifit=0
                self.prepare=0
                self.productivity=0
                self.allow=False          
    class product:
        def __init__(self,productname):
            self.name=productname# 记录名字
            self.benifit={}
            self.thelist=[]#记录能不能种各种田，利润是多少
            self.limit=0#销量预期初始化
            self.nowlimit=0#已经销售量初始化
            self.theamount=[]#计算销量预期所用数据
            self.amountgetter(productname)#得到self.theamount
            for i in self.theamount:
                self.limit+=i[0]*i[1]#销量为每个田的亩数乘以该田亩产
            self.datagetter(productname)#计算每种种植的利润
            #print(self.name,self.limit,'*',self.theamount)
            self.adding()#似乎没用
            self.thelist=sorted(self.thelist,key=lambda x:x.benifit,reverse=True)#按照利润大小排序
        def biggest_one_getter(self):
            #得到最大利润的地块种类，若都不能种返回"NULL"
            num=0
            for i in self.thelist:
                if i.allow:
                    return [num,i.blockname]#
                num+=1 
            return [0,"NULL"]      
        def amountgetter(self,productname):#得到销量限制
            thefirst_list=[]
            i=-1
            for j in range(86):
                if df2.at[j,'作物名称']==productname:#找到所有种的东西是目标的情况
                    i+=1
                    thefirst_list.append([df2.at[j,'种植面积/亩']])#将面积填入
                    if not pd.isna(df2.at[j,'种植地块']):#如果没有种植地块，填入空
                        if df2.at[j,'种植地块'][0]=='A':#辨别地块种类，填入种类
                            thefirst_list[i].append(Plan.blockname_lexico[0])
                        if df2.at[j,'种植地块'][0]=='B':
                            thefirst_list[i].append(Plan.blockname_lexico[1])
                        if df2.at[j,'种植地块'][0]=='C':
                            thefirst_list[i].append(Plan.blockname_lexico[2])
                        if df2.at[j,'种植地块'][0]=='D':
                            thefirst_list[i].append(Plan.blockname_lexico[3])
                        if df2.at[j,'种植地块'][0]=='E':
                            thefirst_list[i].append(Plan.blockname_lexico[4])
                        if df2.at[j,'种植地块'][0]=='F':
                            thefirst_list[i].append(Plan.blockname_lexico[5])
                    else :
                        thepos=self.findposer(j)
                        thefirst_list[i].append(Plan.blockname_lexico[self.changer(df2.at[thepos,'种植地块'][0])])
            for i in range(len(thefirst_list)):
                for j in range(106):#在表格三中找到对应的亩产量
                    if thefirst_list[i][1]==df3.at[j,'地块类型'] and productname==df3.at[j,'作物名称']:
                        self.theamount.append([thefirst_list[i][0],df3.at[j,'亩产量/斤']])#记录亩产和面积
                        break
        def findposer(self,i):
            while True:
                i-=1
                if not pd.isna(df2.at[i,'种植地块']):
                    return i
        def changer(self,char):
            if char=='A':return 0
            if char=='B':return 1
            if char=='C':return 2
            if char=='D':return 3
            if char=='E':return 4
            if char=='F':return 5
        def datagetter(self,productname):#将该作物种在各种地块利润进行填写
            for i in range(6):#初始化说有地块无法种植，利润为0
                self.thelist.append(Plan.datatype(Plan.blockname_lexico[i],0,0,0,0,False))
            for i in range(106):#遍历数据集
                if df3.at[i,'作物名称']==productname:#找到了该作物
                   index=Plan.block_reverse[df3.at[i,'地块类型']]#得到其在block_lexico中的位置
                   rrange=df3.at[i,'销售单价/(元/斤)']#得到价格
                   box=rrange.split('-')
                   low=float(box[0])#最低价
                   high=float(box[1])#最高价
                   p1=float(df3.at[i,'种植成本/(元/亩)'])
                   p2=float(df3.at[i,'亩产量/斤'])
                   theitem=Plan.datatype(Plan.blockname_lexico[index],low,high,p1,p2,True)#建立对象
                   self.thelist[index]=theitem#将对象填入
        def adding(self):#似乎没有什么用
            for i in self.thelist:
                self.benifit[i.blockname]=i.benifit
    # 程序的主体部分
    def filingit(self):
    # 该函数用于对dou_lexico,liang_lexico,jun_lexico,cai_lexico进行初始化
        for i in range(86):
            Plan.dou_lexico[df2.at[i,'作物名称']]=0
            Plan.liang_lexico[df2.at[i,'作物名称']]=0
            Plan.jun_lexico[df2.at[i,'作物名称']]=0
            Plan.cai_lexico[df2.at[i,'作物名称']]=0
        self.subfile("粮食")
        self.subfile("豆类")
        self.subfile("蔬菜 ")
        self.subfile("食用菌")
    def subfile(self,name):
        for i in range(86):
            if name==Plan.plant_lexico[0]:
                if name==df2.at[i,'作物类型'] or df2.at[i,'作物类型']=="粮食（豆类）":
                   Plan.liang_lexico[df2.at[i,'作物名称']]=1
            if name==Plan.plant_lexico[1]:
                if df2.at[i,'作物类型']=="粮食（豆类）" or df2.at[i,'作物类型']=="蔬菜（豆类）":#将豆写到豆类中，特殊处理
                    Plan.dou_lexico[df2.at[i,'作物名称']]=1
            if name==Plan.plant_lexico[2]:
                if name==df2.at[i,'作物类型'] or df2.at[i,'作物类型']=="蔬菜（豆类）":
                   Plan.cai_lexico[df2.at[i,'作物名称']]=1
            if name==Plan.plant_lexico[3]:
                if name==df2.at[i,'作物类型']:
                   Plan.jun_lexico[df2.at[i,'作物名称']]=1
    def __init__(self,times):#times 表示计算的年份共有多少
        self.ii=0
        self.times=times
        self.filingit()
        self.theproducts=[]#保存作物种类变量（Product类对象）
        self.theblocks={}#保存地块对象（block类对象）
        self.datain()#读取数据并建立结构的函数
        self.stringbox=""#无用
        self.sorting()#事先将作物构成的集合进行排序
        for i in range(times):
            self.working()#计算每一年
            self.ii+=1
        self.player()#将每一年的结果打印出来
    def player(self):
        #打印结果
        self.result=[]
        for i in self.theblocks.keys():
            for j in self.theblocks[i]:
                self.result.append(j.result[0])
        for i in self.theblocks.keys():
            item=self.theblocks[i]
            if item[0].selfname=="水浇地" or item[0].selfname=="普通大棚 " or item[0].selfname=="智慧大棚":
                for j in self.theblocks[i]:
                    self.result.append(j.result[1])
    def season_two(self):
        for product in self.theproducts:
            for i in product.thelist:
                if i.i!=0:
                    if i.blockname=="水浇地" :
                       i.allow=True
                       i.benifit=i.prepare
                    if i.blockname=="普通大棚 ":
                       i.allow=True
                       i.benifit=i.prepare
                    if i.blockname=="智慧大棚":
                       i.beifit=i.prepare
                       i.allow=True 
        for i in self.theblocks.keys():
            for j in self.theblocks[i]:
                if j.status==[2,0] and j.newmemo[0]!=["水稻"]:
                    j.status=[2,1]
                    j.amount*=2
                if j.status==[3,0] or j.status==[4,0]:
                    j.status[1]=1
                    j.amount*=2               
    def working(self):   
        #if self.ii==1:
            #print("*********")     
        while self.teller():#判断还可以获利，继续种
            self.allocated(False)#给地块分配作物
            self.sorting()# 重新排序得到最大收益的种法
        #if self.ii==1:
            #print("********")
        self.season_two()
        self.sorting() 
        while self.teller():
            self.allocated(True)
            self.sorting()
        if self.ii<self.times-1:
            self.reversion(True)#将数据进行调整方便明年的计算
        else:
            self.reversion(False)
    def reversion(self,teller):
        self.result=[]
        for i in self.theblocks.keys():
            for j in self.theblocks[i]:
                j.limittimes+=1 #1
                j.update()      #2
                #入手的1和2联合工作将使本年没种豆子的地年份加一
                j.status[1]=0
                #重新设置成一季都没种
                j.memo=j.newmemo
                j.newmemo=[[],[]]
                #接受新一年方案的量重新清空
                j.nowamount=0
                j.amount=j.trueamount
                #种地量清空
                if teller:
                   j.result=[[],[]]
        for item in self.theproducts:
            item.nowlimit=0
            for i in item.thelist:
                if i.allow==False and i.i!=0: #本来可以种但应为今年的限制没法种的，重新设置为可以种
                    i.allow=True
                    i.benifit=i.prepare
            self.nowlimit=0#本年的实际种植量清空
        self.sorting()
        return
    def finder(self,item):
        for i in item.newmemo[0]:#如果第一季有水稻为真，否则为假
            if i=="水稻":
                return True
        return False
    def bool_teller(self,item):
        teller1=item.status==[3,1] or item.status==[4,1]#大鹏类种了一季度
        teller2=item.status==[2,1] and not(self.finder(item))#水浇地中了一季节且不是水稻
        teller=teller1 or teller2
        return teller        
    def allocated(self,order):
        item=self.theproducts[0]#目前利润最大的作物
        amount=item.limit-item.nowlimit#计算还能种多少
        dataset=item.biggest_one_getter()#得到最大利润的地块种类（准确的说它的第一个才是目标地块）
        aiming_block_list=self.theblocks[dataset[1]]#得到目标地块的列表
        func=lambda x:x.amount-x.nowamount if self.bool_teller(x) else x.trueamount-x.nowamount#判断地块剩余量的函数
        aiming_block_list=sorted(aiming_block_list,key=lambda x:func(x),reverse=True)#优先选择剩余面积大的地块填充
        blocksizeneed=amount/item.thelist[dataset[0]].productivity#计算种植到销量需要多少该种地块
        if blocksizeneed<=0:
            self.modulator(item)
            return
        for i in aiming_block_list:#从大到小遍历地块
            if item.thelist[dataset[0]].allow:#该种作物目前能种在这种地块上
                if i.status[1]==0: #首先讨论大鹏没种的情况
                    if  blocksizeneed<=i.trueamount-i.nowamount:# 能种下
                        timer=i.add_element(item.name,blocksizeneed)#向地块种植东西，返回0表示成功，返回一表示失败
                        if timer==0:#种植失败继续操作
                            continue
                        self.modulator(item)#种植成功，销量用完，不能再种该作物
                        item.nowlimit=item.limit
                        blocksizeneed=0
                        break
                    if blocksizeneed>i.trueamount-i.nowamount:#该地种不玩
                        a=i.trueamount-i.nowamount
                        timer=i.add_element(item.name,i.trueamount-i.nowamount)#种进去
                        if timer==0:#失败下一个
                            continue
                        blocksizeneed-=a#剩余销量的变化
                        item.nowlimit+=a*item.thelist[dataset[0]].productivity#实际种植的量
                if self.bool_teller(i) or i.status[0]==1 or i.status[0]==2:#另一种情况，与上文基本一
                    if blocksizeneed<=i.trueamount*2-i.nowamount:
                        if 0==self.predo(item,i):
                            break
                        a=i.trueamount*2-i.nowamount
                        timer=i.add_element(item.name,a)
                        if timer==0:
                            continue
                        self.modulator(item)
                        item.nowlimit=item.limit
                        blocksizeneed=0
                        break
                    if blocksizeneed>i.trueamount*2-i.nowamount:
                        if 0==self.predo(item,i):
                            continue
                        a=i.trueamount*2-i.nowamount
                        timer=i.add_element(item.name,a)
                        if timer==0:
                            continue
                        blocksizeneed-=a
                        item.nowlimit+=a*item.thelist[dataset[0]].productivity
        item.thelist[dataset[0]].allow=False#该种作物不能再种在此种地上了
    def predo(self,product,block):
        if block.status==[3,1]:
            if len(block.newmemo[0])!=0 and Plan.cai_lexico[block.newmemo[0][0]]==1:
                if Plan.jun_lexico[product.name]!=1:
                   return 0
            else:
                if len(block.newmemo[0])!=0 and Plan.jun_lexico[block.newmemo[0][0]]==1:
                    if Plan.cai_lexico[product.name]!=1:
                        return 0
        return 1
    def modulator(self,item):#不能再种某作物
        for i in item.thelist:
            i.allow=False                    
    def teller(self):
        for i in self.theproducts:#遍历每种作物
            if i.biggest_one_getter()[1]!="NULL":#如果还有种的
                return True#返回能种信息
            return False#返回不能种信息
    def sorting(self):
        equal_func0=lambda x:x.biggest_one_getter()#得到种某作物最大利润的地块种类
        equal_func1=lambda x:0 if equal_func0(x)[1]=="NULL" else x.thelist[equal_func0(x)[0]].benifit 
        #若没有能种的地块种类，返回0，否则返回利润，依靠此排序
        self.theproducts=sorted(self.theproducts,key=equal_func1,reverse=True)        
    def shifter(self,alg):
        #计算工具，用地块编号的前缀得到地块种类
        if alg=='A':return Plan.blockname_lexico[0]
        if alg=='B':return Plan.blockname_lexico[1]
        if alg=='C':return Plan.blockname_lexico[2]
        if alg=='D':return Plan.blockname_lexico[3]
        if alg=='E':return Plan.blockname_lexico[4]
        if alg=='F':return Plan.blockname_lexico[5]
    def datain(self):# 得到数据
        for i in range(6):
            self.theblocks[Plan.blockname_lexico[i]]=[]
        #为每种地块建立一个集合
        theprepare=[]#保存数据用
        thelist=[[],[]]#保存数据用
        for i in range(86):#遍历整个文件
            if  not pd.isna(df2.at[i,'种植地块']):#上一个地块数据收集完毕
                if i!=0:#不是第一个
                    self.theblocks[theprepare[0]].append(Plan.block(theprepare[0],thelist,theprepare[1]))#将数据填入
                theprepare=[]#清空数据
                thelist=[[],[]]#清空数据
                theprepare=[self.shifter(df2.at[i,'种植地块'][0]),float(df2.at[i,'种植面积/亩'])]#重新开始装数据
                thelist[0].append(df2.at[i,'作物名称'])#重新开始装数据
                continue
            if i==85:#都最后一行，特殊处理
                thelist[1].append(df2.at[i,'作物名称'])#将种植的植物装入相关数据集
                self.theblocks[theprepare[0]].append(Plan.block(theprepare[0],thelist,theprepare[1]))#将完成的地块信息填入
                continue
            if  not pd.isna(df2.at[i,'种植地块']):#没有完成地块信息
                if df2.at[i,'种植季次']=='单季' or df2.at[i,'种植季次']=='第一季':
                    thelist[0].append(df2.at[i,'作物名称'])#第一季装入第一季的数据集
                    theprepare[1]+=float(df2.at[i,'种植面积/亩'])#计算地块的总面积
                else:
                    thelist[1].append(df2.at[i,'作物名称'])#第二季装入第二季的数据集
                    theprepare[1]+=float(df2.at[i,'种植面积/亩'])#计算地块的总面积
        theformat={}
        for i in range(86):
            theformat[df2.at[i,'作物名称']]=0#每种小类作物占一个位置
        for i in range(86):
            if theformat[df2.at[i,'作物名称']]==0:#对于第一遍访问到的小作物，建立对象
                theformat[df2.at[i,'作物名称']]=1
                self.theproducts.append(Plan.product(df2.at[i,'作物名称']))
def finder(name,sheet):
    i=1
    for i in range(1,100):
        if sheet.cell(row=1,column=i).value==name:
            return i
workbook=openpyxl.load_workbook('D:\\programming\\PYCD\\result1_1.xlsx')
working=[]
working.append(workbook['2024'])
working.append(workbook['2025'])
working.append(workbook['2026'])
working.append(workbook['2027'])
working.append(workbook['2028'])
working.append(workbook['2029'])
working.append(workbook['2030'])
i=0
doit=Plan(1)
for sheet in working:
    i+=1
    doit=Plan(i)
    for row in range(len(doit.result)):
        for size in range(len(doit.result[row])):
            colnum=finder(doit.result[row][size][0],sheet)
            sheet.cell(row=row+2,column=colnum).value=doit.result[row][size][1]
            warnings.warn("Cannot parse header or footer so it will be ignored")
            workbook.save('D:\\programming\\PYCD\\result1_1_modified.xlsx')                
